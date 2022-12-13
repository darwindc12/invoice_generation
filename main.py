from fpdf import FPDF
import pandas
import glob
import openpyxl
import pathlib

filepaths = glob.glob("invoice/*.xlsx")


for filepath in filepaths:

    pdf = FPDF(orientation="P", unit='mm', format="A4")
    pdf.add_page()
    filename = pathlib.Path(filepath).stem
    nv, date = filename.split("-")

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Invoice nr. {nv}", align="L", ln=1)

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Date {date}", align="L", ln=1)

    df = pandas.read_excel(filepath, sheet_name="Sheet 1")

    column_name = [item.replace("_", " ").title() for item in df]
    pdf.set_font(family="Times", style="B", size=12)
    pdf.cell(w=30, h=8, txt=column_name[0], border=1)
    pdf.cell(w=50, h=8, txt=column_name[1], border=1)
    pdf.cell(w=40, h=8, txt=column_name[2], border=1)
    pdf.cell(w=30, h=8, txt=column_name[3], border=1)
    pdf.cell(w=30, h=8, txt=column_name[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", style="B", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=50, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=40, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    pdf.output(f"PDFs/{filename}.pdf")

